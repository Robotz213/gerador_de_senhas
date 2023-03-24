from PySide2.QtWidgets import(QApplication, QWidget)
from PyQt5 import QtWidgets
import os
import sys
import openpyxl
from Scripts.ui_uploadfiles import Ui_Upload
import os
import openpyxl
from pathlib import *
import sys
from resources import BotResources
from tkinter.messagebox import showinfo
import string
import pathlib
import random
import shutil



class Uploadfiles(QWidget, Ui_Upload):

    def __init__(self) -> None:
        super(Uploadfiles, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Gerador de Senhas Em Lotes")
        self.BuscarPlanilha.clicked.connect(lambda: self.openfile())

        self.IniciarUpload.clicked.connect(lambda: self.initbot(input_filename=self.lineEndr.text()))
    
    def openfile(self):
        input_filename = QtWidgets.QFileDialog.getOpenFileName()[0]
        self.lineEndr.setText(input_filename)
                

    def initbot(self, input_filename):
            

            self.p = input_filename
            wrkbk_input = openpyxl.load_workbook(filename=self.p)
            sheet_input = wrkbk_input.active
            
            self.output_dir_path = pathlib.Path(self.p).parent.resolve()
            self.copyfile = os.path.join(pathlib.Path(__file__).parent.resolve(), 'template', 'output.xlsx')
            self.output_filename = '{dir_path}/output.xlsx'.format(dir_path=str(self.output_dir_path))
            shutil.copy(self.copyfile,self.output_filename)



            for i in range(2, sheet_input.max_row+1):
                    cell_obj = sheet_input.cell(row=i, column=1)
                    if cell_obj.value is not None and cell_obj.value != '':
                        cpf = sheet_input.cell(row=i, column=3).value
                        self.row= i-1
                        self.cpf = cpf
                        info = [cpf]
                        self.generatesenha(info)
        
                        if i == sheet_input.max_row:
                            showinfo('SUCESSO', 'SENHAS GERADAS')
            

    def generatesenha(self, info):

        password = ''.join(random.choices(string.ascii_uppercase + string.digits + string.ascii_lowercase, k=8))
        self.append_senhas(password)
    
    def append_senhas(self, password):
        

        sheet = 'Sheet1'
        append_data = [self.cpf,password]
        result_append_information_on_output = BotResources().append_information_on_output(
            output_path=self.output_filename,
            data=append_data,
            sheet=sheet
        )
        
        if result_append_information_on_output is True:
            print("Senha {} criada com sucesso!".format(self.row))
        else:
            print(Exception)


if __name__ == "__main__":
    
    
    app = QApplication(sys.argv)
    window = Uploadfiles()
    window.show()
    app.exec_()
    