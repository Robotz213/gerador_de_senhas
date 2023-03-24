from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

from time import sleep
from termcolor import colored
from printlog import print_log as prt

import subprocess
import openpyxl
import shutil
import sys
import os
import string

class BotResources:
    def __init__(self):
        self.wait_time = 10
    
    def open_output(self, dir_path, spreadsheet_path):
        prt(type='log', message='Fim dos processos')
        try:
            if sys.platform == 'linux':
                os.system('xdg-open "%s"' % dir_path)
                subprocess.call(["xdg-open", spreadsheet_path])
            else:
                os.system(f'start "{dir_path}"')
                os.startfile(spreadsheet_path)
                os.system("pause")
            
            return True
        except Exception as e:
            prt(type='error', message=e)

            return False

    def clear_output(self, file_base, file_copy):
        try:
            shutil.copyfile(file_base, file_copy)
            output_filename = file_copy
            wrkbk_output = openpyxl.load_workbook(filename=output_filename)
            sheet_output = wrkbk_output.active
            sheet_output.delete_rows(2, sheet_output.max_row+1)
            wrkbk_output.save(output_filename)

            return True
        except Exception as e:
            print(e)

            return False

    def append_information_on_output(self, output_path, data, sheet = 'Sheet1'):
        try:
            wb = openpyxl.load_workbook(filename=output_path)
            sheet = wb[sheet]

            sheet.append(data)
            wb.save(output_path)

            return True
        except Exception as e:
            print(e)

            return e

    def login_esaj(self, credentials, loginmethod, driver):

        if loginmethod == "certificado":
            
            certinfo = ''.join(filter(lambda x: x not in string.punctuation, credentials['login']))
            print(certinfo)
            driver.get("https://consultasaj.tjam.jus.br/sajcas/login#aba-certificado")
            sleep(3)
            logincert = WebDriverWait(driver, self.wait_time).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="certificados"]')))
            sleep(3)

            for cert in logincert:
                loginopt = cert.find_elements(By.TAG_NAME, "option")
                
                for option in loginopt:
                    

                        if  certinfo  in option.text.lower():
                            try:
                                sencert = option.get_attribute("value")
                                select = Select(driver.find_element(By.CSS_SELECTOR, 'select[id="certificados"]'))
                                select.select_by_value(sencert)
                                entrar = driver.find_element(By.XPATH, '//*[@id="submitCertificado"]')
                                entrar.click()
                                prt(type="log", message="confirme o websigner", row=1)
                                sleep(3)
                                prt(type='log', message='Login efetuado com sucesso', row=1)
                                sleep(1)

                                return True
                            except Exception as e:
                                print(e)

                                return False

        elif loginmethod == "cpf":
            try:
                driver.get("https://consultasaj.tjam.jus.br/sajcas/login")
                sleep(3)

                userlogin = driver.find_element(By. CSS_SELECTOR, '#usernameForm')
                userlogin.click()
                userlogin.send_keys(credentials['login'])

                userpass = driver.find_element(By. CSS_SELECTOR, '#passwordForm')
                userpass.click()
                userpass.send_keys(credentials['password'])
                entrar = driver.find_element(By.CSS_SELECTOR, '#pbEntrar')
                entrar.click()
                sleep(2)

                return True
            except Exception as e:
                print(e)

                return False

    def copy_files_log(self, file_base, file_copy):

        shutil.copyfile(file_base, file_copy)